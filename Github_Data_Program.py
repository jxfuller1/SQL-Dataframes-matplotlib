
from PyQt5.QtCore import Qt, pyqtSignal, QRect, QThread
from PyQt5.QtGui import QCursor, QFont, QPixmap
from PyQt5.QtWidgets import QWidget, QStackedLayout, QLabel, QVBoxLayout, QLineEdit, QHBoxLayout, QPushButton, \
    QListView, QGroupBox, QRadioButton, QFrame, QListWidget, QScrollArea, QListWidgetItem, QMessageBox, QCheckBox, \
    QComboBox, QTableWidget, QTableWidgetItem, QApplication, QHeaderView
import sqlite3
import pandas as pd
import time
import connectorx as cx
import urllib
import os
from datetime import datetime
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import calendar
import getpass
import win32print
import win32com.client

from PIL import Image
from openpyxl.drawing.image import Image as Im
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU

# get current working directory (needed if using on different computers)
my_path = os.path.abspath(os.path.dirname(__file__))

# add database connection, for connectorx module for sqlite database
db_path = urllib.parse.quote('your database')
connect = 'sqlite://' + db_path


# =========================================================================================================
# =================================== Tom's Filter's for program===========================================
# =========================================================================================================
# Tom's Filter's for program
# substrings to check for against string later
workcenter_removal = ["DBA", "INSP", "TOOL", "MRB", "TEST LAB", "TEST PREP", "PP-ADMIN", "SHIP/REC", "APEX", "TEST",
                      "TRAINING", "ASTRONICS"]

# substrings to check for against string later
process_id_removal = ["TRANSFER", "ISSUE", "TOOL", "INSPECT"]

# check to make sure string contains "JOB"
jobnumber_test = ["JOB", "FAI"]

ncr_columns_display = ["Date", "NCR No", "Job No/PO", "Drawing No", "Title", "Failure Category", "Disposition", "Close Date"]

# =========================================================================================================
# =========================================================================================================
# =========================================================================================================


class Program_layout(QStackedLayout):
    # any signals between program and UI go here
    resize_window = pyqtSignal(str)
    print_selection = pyqtSignal(str)

    # signal for onsizeChanged function for UI_Template (must be here, unless this function is removed from Main UI File)
   # sizeChanged = pyqtSignal(int, int)

    # add any other signals between program and UI go here

    def __init__(self, parent):
        super(Program_layout, self).__init__(parent)

        # put charts for user to choose from in a list for combobox in UI
        self.charts_available = ["PK JOBs", "AK JOBs", "FA JOBs", "SA JOBs", "FA SEQs", "Overall JOBs",
                                 "PO Line Items", "Total PO Items"]

        # initialize chart_months var, capturing the months chosen in the UI at the time the load charts button is hit
        # which will be used later for populating the graphs
        self.chart_months = []

        # values for charts when data is obtained, this is needed so that i don' thave to reload the data
        # from the database when changing chart type
        self.job_data = {}
        self.ncr_data = {}
        self.qc_data = {}
        self.fa_ncr_data = []
        self.sa_ncr_data = []
        self.pk_ncr_data = []
        self.ak_ncr_data = []
        self.total_jobs_ncr_data = []
        self.po_ncr_data = []

        self.program_mainwork_UI()

    # call each page of UI
    def program_mainwork_UI(self):
        self.home_page()
        self.charts_page()
        self.add_month_data_page()
        self.data_filter_page()
        self.options_page()
        self.vendor_filters()
        self.about_page()

        # run sql query class for caching default queries  , pass the last 12 items in the combobox
        # this will be used for what tables to query in the database
        last_12_items = [self.comboBox_start.itemText(i) for i in range(max(self.comboBox_start.count() - 12, 0), self.comboBox_start.count())]

        # need job items as well to pass to sql query caching
        job_items = [self.list_type.item(i).text() for i in range(self.list_type.count()) if
                     self.list_type.item(i).checkState() == 2]

        self.sql_cached = sql_cache(last_12_items, job_items)
        self.sql_cached.start()

    # home page of UI
    def home_page(self):
        self.home_scroll = QScrollArea()
        self.home_scroll.setWidgetResizable(True)
        self.home_widget = QWidget()

        self.home_layout = QVBoxLayout()
        image_path = my_path + "\\plane.PNG"
        pixmap = QPixmap(image_path)

        # Create a QLabel and set its properties
        label = QLabel()
        label.setPixmap(pixmap.scaled(480, 200))
        label.setAlignment(Qt.AlignCenter)
        self.home_layout.addWidget(label)

        self.home_widget.setLayout(self.home_layout)

        self.home_scroll.setWidget(self.home_widget)

        self.addWidget(self.home_scroll)

# =============================chart page===========================================================
# =============================chart page===========================================================
# =============================chart page===========================================================
    # chart page UI setup
    def charts_page(self):

        self.charts_scroll = QScrollArea()
        self.charts_scroll.setWidgetResizable(True)
        self.charts_widget = QWidget()

        self.verticalLayout_3 = QVBoxLayout()

        myfont = QFont()
        myfont.setPointSize(12)
        self.label = QLabel("<b>Charts</b>")
        self.label.setAlignment(Qt.AlignHCenter)
        self.label.setFont(myfont)

        self.verticalLayout_3.addWidget(self.label)
        self.verticalLayout_3.addSpacing(20)

        self.horizontalLayout_2 = QHBoxLayout()
        self.horizontalLayout_2.addStretch()

        myfont.setPointSize(8)
        self.label_2 = QLabel("<b>Start Month/Year</b>")
        self.label_2.setFont(myfont)

        self.horizontalLayout_2.addWidget(self.label_2)

        self.comboBox_start = QComboBox()

        self.horizontalLayout_2.addWidget(self.comboBox_start)
        self.horizontalLayout_2.addSpacing(10)

        self.label_3 = QLabel("<b>End Month/Year</b>")
        self.label_3.setFont(myfont)
        self.horizontalLayout_2.addWidget(self.label_3)

        self.comboBox_end = QComboBox()

        self.horizontalLayout_2.addWidget(self.comboBox_end)
        self.horizontalLayout_2.addStretch()

        self.pushButton = QPushButton("Load Charts")
        self.pushButton.setToolTip("Charts will be based on filters chosen from Filters!")
        self.pushButton.clicked.connect(self.charts_check_month)

        self.horizontalLayout_2.addWidget(self.pushButton)

        self.pushButton_print = QPushButton("Print")
        self.pushButton_print.setToolTip("Note: The printed Charts will reflect how they look in this window!")
        self.pushButton_print.clicked.connect(self.setup_print_charts)
        self.pushButton_print.setEnabled(False)

        self.horizontalLayout_2.addWidget(self.pushButton_print)

        self.verticalLayout_3.addLayout(self.horizontalLayout_2)
        self.verticalLayout_3.addSpacing(10)

        self.line = QFrame()
        self.line.setFrameShape(QFrame.HLine)
        self.line.setFrameShadow(QFrame.Sunken)
        self.line.setStyleSheet("border: 2px solid black;")

        self.verticalLayout_3.addWidget(self.line)
        self.verticalLayout_3.addSpacing(10)

        self.combo_chart_layout = QHBoxLayout()
        self.combo_chart_label = QLabel("<b>Choose Chart:</b>")
        self.combo_chart_label.setFont(myfont)

        self.combo_chart = QComboBox()
        self.combo_chart.setEnabled(False)
        self.combo_chart.addItems(self.charts_available)
        self.combo_chart.currentIndexChanged.connect(self.on_chart_change)

        self.combo_chart_layout.addStretch()
        self.combo_chart_layout.addWidget(self.combo_chart_label)
        self.combo_chart_layout.addWidget(self.combo_chart)
        self.combo_chart_layout.addStretch()
        self.verticalLayout_3.addLayout(self.combo_chart_layout)

        self.horizontalLayout = QHBoxLayout()

        self.verticalLayout = QVBoxLayout()

        myfont.setPointSize(10)
        self.label_4 = QLabel("<b>Month Chart</b>")
        self.label_4.setFont(myfont)
        self.label_4.setAlignment(Qt.AlignHCenter)

        self.verticalLayout.addWidget(self.label_4)

        # a figure instance to plot on
        self.figure_12_month = plt.figure(tight_layout=True)

        # for font size
        self.ax_12_month = self.figure_12_month.add_subplot(111)
        self.ax_12_month.tick_params(axis='both', labelsize=7)
        self.ax_12_month.tick_params(axis='x', rotation=90)

        # this is the Canvas Widget that
        # displays the 'figure'it takes the
        # 'figure' instance as a parameter to __init__
        self.canvas_12_month = FigureCanvas(self.figure_12_month)

        # this is the Navigation widget
        # it takes the Canvas widget and a parent
        self.toolbar_12_month = NavigationToolbar(self.canvas_12_month)
        self.toolbar_12_month.setMaximumHeight(22)

        self.verticalLayout.addWidget(self.toolbar_12_month)
        self.verticalLayout.addWidget(self.canvas_12_month)

        self.horizontalLayout.addLayout(self.verticalLayout)

        self.verticalLayout_2 = QVBoxLayout()

        self.label_5 = QLabel("<b>Failure Categories</b>")
        self.label_5.setFont(myfont)
        self.label_5.setAlignment(Qt.AlignHCenter)

        self.verticalLayout_2.addWidget(self.label_5)

        # a figure instance to plot on
        self.figure_lastmonth = plt.figure(tight_layout=True)

        # set basic graph for for UI sake... this all gets replaced when data gets called for graph
        self.ax_lastmonth = self.figure_lastmonth.add_subplot(111)
        self.ax_lastmonth.tick_params(axis='both', labelsize=7)
        self.ax_lastmonth.tick_params(axis='x', rotation=90)

        # this is the Canvas Widget that
        # displays the 'figure'it takes the
        # 'figure' instance as a parameter to __init__
        self.canvas_lastmonth = FigureCanvas(self.figure_lastmonth)

        # this is the Navigation widget
        # it takes the Canvas widget and a parent
        self.toolbar_lastmonth = NavigationToolbar(self.canvas_lastmonth)
        self.toolbar_lastmonth.setMaximumHeight(22)

        self.verticalLayout_2.addWidget(self.toolbar_lastmonth)
        self.verticalLayout_2.addWidget(self.canvas_lastmonth)

        self.horizontalLayout.addLayout(self.verticalLayout_2)

        self.verticalLayout_3.addLayout(self.horizontalLayout)
        self.verticalLayout_3.addSpacing(20)

        self.line_2 = QFrame()
        self.line_2.setFrameShape(QFrame.HLine)
        self.line_2.setFrameShadow(QFrame.Sunken)

        self.verticalLayout_3.addWidget(self.line_2)
       # self.verticalLayout_3.addSpacing(20)

        self.label_6 = QLabel("<b>Month NCR DATA</b>")
        self.label_6.setAlignment(Qt.AlignHCenter)
        self.label_6.setFont(myfont)

        self.verticalLayout_3.addWidget(self.label_6)

        self.tableWidget = TableWithCopy()
        # make cells stretch to width of UI for better look
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # alternate row colors for better lookign table
        self.tableWidget.setAlternatingRowColors(True)
        self.tableWidget.horizontalHeader().setSectionsMovable(True)
        self.tableWidget.setMaximumHeight(135)

        self.verticalLayout_3.addWidget(self.tableWidget)

        self.verticalLayout_3.setStretch(4, 9)
        self.verticalLayout_3.setStretch(7, 1)

        self.charts_widget.setLayout(self.verticalLayout_3)

        self.charts_scroll.setWidget(self.charts_widget)

        self.addWidget(self.charts_scroll)


        # get all job tables, this will be for putting into comboboxes for user to choose
        self.months_ordered = self.charts_get_tables("JOB")

        # remove the "JOB" from the lists for the comboboxes
        months = [item.replace("JOB", "") for item in self.months_ordered]

        # replace the last digit with the corresponding calender name for the combobox
        months = [x.split()[0] + " " + calendar.month_name[int(x.split()[1])] for x in months]

        # add months to combobox
        self.comboBox_start.addItems(months)
        self.comboBox_end.addItems(months)

    # get tables from database
    def charts_get_tables(self, table_type):
        conn = sqlite3.connect("your database")
        cursor = conn.cursor()

        query = "SELECT name FROM sqlite_master WHERE type='table'"
        cursor.execute(query)

        # get all table names
        fetch = cursor.fetchall()

        conn.close()

        # pass table names into a list instead of tuples
        alltable_list = [item for tpl in fetch for item in tpl]

        # reduce list to just table_type
        table_list = [item for item in alltable_list if table_type in item]

        # convert tables names purely to integers for ordering them correctly
        ordered_tables_raw = []
        for i in table_list:
            second_value = i.split(" ")[1]
            if len(second_value) == 1:
                second_value = "0" + i.split(" ")[1]

            ordered_tables_raw.append(int(i.split(" ")[0] + second_value))

        # order table after conversion to integers
        ordered_tables_raw.sort()

        # convert the ordered integer table back to ordered list of table names
        ordered_tables_clean = []
        for i in ordered_tables_raw:
            # get year
            year = str(i)[:4]

            # get month, but replace 0 if on front of it
            month = str(i)[4:]
            if month[0] == "0":
                month = month.replace("0", "")

            ordered_tables_clean.append(year + " " + month + " " + table_type)

        return ordered_tables_clean

    # get data from UI and start Qthread to collect data
    def charts_check_month(self):
        # collect data from UI JOB filters for passing to Qthread DATA
        workcenter_items = [self.list_workcenter.item(i).text() for i in range(self.list_workcenter.count()) if
                            self.list_workcenter.item(i).checkState() == 0]

        process_items = [self.list_process.item(i).text() for i in range(self.list_process.count()) if
                         self.list_process.item(i).checkState() == 0]

        job_items = [self.list_type.item(i).text() for i in range(self.list_type.count()) if
                     self.list_type.item(i).checkState() == 2]

        fa_seq_state = self.checkBox_fa.checkState()


        # collect data from UI NCR filters for passing to Qthread DATA
        ncr_columns = [self.list_ncr_data.item(i).text() for i in range(self.list_ncr_data.count()) if
                       self.list_ncr_data.item(i).checkState() == 2]

        ncr_drawing = self.checkBox_drawing.checkState()
        ncr_pks = self.checkBox_linked.checkState()
        ncr_novendorpart = self.checkBox_vendor.checkState()


        # collect data from UI Vendor filters for passing to Qthread DATA
        vendor_items = [self.listWidget_vendors.item(i).text() for i in range(self.listWidget_vendors.count()) if
                        self.listWidget_vendors.item(i).checkState() == 0]

        # this is to make sure someone isn't dumb and wants to remove every entry , otherwise program will crash
        sanity_check = False
        if len(job_items) != 0:
            sanity_check = True

        # check to make sure end chart has a higher index number for month range
        if self.comboBox_start.currentIndex() <= self.comboBox_end.currentIndex() and sanity_check == True:
            # get list of table names from combobox months index chosen based on the self.months_ordered variable which
            # gets the table names from the database; these will be the table names to parse to collect data
            self.job_month_range = self.months_ordered[self.comboBox_start.currentIndex():self.comboBox_end.currentIndex()+1]

            # get current selected months and pass to top level var , to be used for graphs later
            self.chart_months = [self.comboBox_start.itemText(i) for i in range(self.comboBox_start.currentIndex(), self.comboBox_end.currentIndex()+1)]

            # replace the "JOB" from the list for the list of ncr tables to parse from database
            self.ncr_month_range = [item.replace("JOB", "PO") for item in self.job_month_range]

            # replace the "JOB" from the list for the list of QC tables to parse from database
            self.qc_month_range = [item.replace("JOB", "QC") for item in self.job_month_range]

            # create Qthread to collect the data from the database
            self.get_graph_data = DATA(self.job_month_range, self.ncr_month_range, self.qc_month_range,
                                       workcenter_items, process_items, job_items, fa_seq_state, ncr_columns,
                                       ncr_drawing, ncr_pks, ncr_novendorpart, vendor_items)
            self.get_graph_data.data_changed.connect(self.on_data_changed)
            self.get_graph_data.start()


        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Either End month earlier than start month. \n "
                        "Or you are removing every entry possible due to custom filters")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec()

    # change charts based on received data from Qthread
    def on_data_changed(self, job_data, ncr_data, qc_data, fa_ncr_data, sa_ncr_data, pk_ncr_data,
                        ak_ncr_data, total_job_ncr_data, po_ncr_data):
        # enable combobox for user to change charts
        self.combo_chart.setEnabled(True)
        self.pushButton_print.setEnabled(True)

        # pass the values to the upper level part of this class, so that the on_chart_change function can use it
        # need to do it this way as the combobox will call the on_chart_change function everytime, have to be done this way
        # otherwise the combobox would reload the data from the database everytime.
        self.job_data = job_data
        self.ncr_data = ncr_data
        self.qc_data = qc_data
        self.fa_ncr_data = fa_ncr_data
        self.sa_ncr_data = sa_ncr_data
        self.pk_ncr_data = pk_ncr_data
        self.ak_ncr_data = ak_ncr_data
        self.total_jobs_ncr_data = total_job_ncr_data
        self.po_ncr_data = po_ncr_data

        # pass info to the function to change the charts
        self.on_chart_change()

    # execute when combobox changes for loadup of charts
    def on_chart_change(self):
        # get last month from the loaded data
        current_monthyear = self.chart_months[-1].split()[1] + " " + self.chart_months[-1].split()[0]

        # these on here as reference for what's in the dicts be passed in
        # job_dict = {"FA_TOTAL": [], "FA_JOBS": [], "SA_JOBS": [], "AKHK_JOBS": [], "PK_JOBS": [], "TOTAL_JOBS": []}

        # ncr_dict = {"FA_JOB_NCRS": [], "SA_JOB_NCRS": [], "PK_JOB_NCRS": [], "AK_JOB_NCRS": [], "TOTAL_JOB_NCRS": [],
        #           "TOTAL_PO_NCRS": []}

        # qc_dict = {"PO_LINE_ITEMS": [], "TOTAL_RECEIVED": []}

        # print(self.job_data)
        # print(self.ncr_data)
        # print(self.qc_data)

        # update Labels on UI for the graphs and set variables to pass to update_graphs to display graphs
        if self.combo_chart.currentText() == "PK JOBs":
            # pass info to 1st graph
            left_y_label = "Count by Month"
            right_y_label = "% JOBSs Non-Conforming"
            title = "% PK JOBs impacted by NCRs by Month"
            bar_legend = ["PK JOBs", "PK NCRs"]
            line_legend = ["% PKs Impacted by NCRs by Month"]
            y_bar1 = self.job_data["PK_JOBS"]
            y_bar2 = self.ncr_data["PK_JOB_NCRS"]
            # get percentages
            y_line1 = [round((b / a) * 100, 1) for a, b in zip(y_bar1, y_bar2)]
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # pass into to 2nd chart
            left_y_label2 = "Count of Failure Categories"
            title2 = "Pareto of Top PK Failure Categories\n" + current_monthyear
            bar_legend2 = [current_monthyear]

            # get categories and values for each
            categories, y_bar2 = self.get_graph_categories(self.pk_ncr_data)
            # update graph
            self.update_lastmonth_graph(categories, y_bar2, left_y_label2, title2, bar_legend2)

            # update table
            self.update_table(self.pk_ncr_data)

        if self.combo_chart.currentText() == "AK JOBs":
            left_y_label = "Count by Month"
            right_y_label = "% JOBSs Non-Conforming"
            title = "% AK JOBs impacted by NCRs by Month"
            bar_legend = ["AK JOBs", "AK NCRs"]
            line_legend = ["% AKs Impacted by NCRs by Month"]
            y_bar1 = self.job_data["AKHK_JOBS"]
            y_bar2 = self.ncr_data["AK_JOB_NCRS"]
            # get percentages
            y_line1 = [round((b / a) * 100, 1) for a, b in zip(y_bar1, y_bar2)]
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # pass into to 2nd chart
            left_y_label2 = "Count of Failure Categories"
            title2 = "Pareto of Top AK Failure Categories\n" + current_monthyear
            bar_legend2 = [current_monthyear]

            # get categories and values for each
            categories, y_bar2 = self.get_graph_categories(self.ak_ncr_data)
            # update graph
            self.update_lastmonth_graph(categories, y_bar2, left_y_label2, title2, bar_legend2)

            # update table
            self.update_table(self.ak_ncr_data)

        if self.combo_chart.currentText() == "FA JOBs":
            left_y_label = "Count by Month"
            right_y_label = "% JOBSs Non-Conforming"
            title = "% FA JOBs impacted by NCRs by Month"
            bar_legend = ["FA JOBs", "FA NCRs"]
            line_legend = ["% FAs Impacted by NCRs by Month"]
            y_bar1 = self.job_data["FA_JOBS"]
            y_bar2 = self.ncr_data["FA_JOB_NCRS"]
            # get percentages
            y_line1 = [round((b / a) * 100, 1) for a, b in zip(y_bar1, y_bar2)]
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # pass into to 2nd chart
            left_y_label2 = "Count of Failure Categories"
            title2 = "Pareto of Top FA Failure Categories\n" + current_monthyear
            bar_legend2 = [current_monthyear]

            # get categories and values for each
            categories, y_bar2 = self.get_graph_categories(self.fa_ncr_data)
            # update graph
            self.update_lastmonth_graph(categories, y_bar2, left_y_label2, title2, bar_legend2)

            # update table
            self.update_table(self.fa_ncr_data)

        if self.combo_chart.currentText() == "SA JOBs":
            left_y_label = "Count by Month"
            right_y_label = "% JOBSs Non-Conforming"
            title = "% SA JOBs impacted by NCRs by Month"
            bar_legend = ["SA JOBs", "SA NCRs"]
            line_legend = ["% SAs Impacted by NCRs by Month"]
            y_bar1 = self.job_data["SA_JOBS"]
            y_bar2 = self.ncr_data["SA_JOB_NCRS"]
            # get percentages
            y_line1 = [round((b / a) * 100, 1) for a, b in zip(y_bar1, y_bar2)]
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # pass into to 2nd chart
            left_y_label2 = "Count of Failure Categories"
            title2 = "Pareto of Top SA Failure Categories\n" + current_monthyear
            bar_legend2 = [current_monthyear]

            # get categories and values for each
            categories, y_bar2 = self.get_graph_categories(self.sa_ncr_data)
            # update graph
            self.update_lastmonth_graph(categories, y_bar2, left_y_label2, title2, bar_legend2)

            # update table
            self.update_table(self.sa_ncr_data)

        if self.combo_chart.currentText() == "FA SEQs":
            left_y_label = "Count by Month"
            right_y_label = ""
            title = "FA Sequences and NCRs by Month"
            bar_legend = ["FA SEQs", "FA NCRs"]
            line_legend = ["% FA Seq. Impacted by NCRs by Month"]
            y_bar1 = self.job_data["FA_TOTAL"]
            y_bar2 = self.ncr_data["FA_JOB_NCRS"]
            # get percentages
            y_line1 = [round((b / a) * 100, 1) for a, b in zip(y_bar1, y_bar2)]
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # pass into to 2nd chart
            left_y_label2 = "Count of Failure Categories"
            title2 = "Pareto of Top FA Failure Categories\n" + current_monthyear
            bar_legend2 = [current_monthyear]

            # get categories and values for each
            categories, y_bar2 = self.get_graph_categories(self.fa_ncr_data)
            # update graph
            self.update_lastmonth_graph(categories, y_bar2, left_y_label2, title2, bar_legend2)

            # update table
            self.update_table(self.fa_ncr_data)

        if self.combo_chart.currentText() == "Overall JOBs":
            left_y_label = "Count by Month"
            right_y_label = "% JOBSs Non-Conforming"
            title = "% JOBs impacted by NCRs by Month"
            bar_legend = ["Overall JOBs", "Overall NCRs"]
            line_legend = ["% JOBs Impact of NCRs by Month"]
            y_bar1 = self.job_data["TOTAL_JOBS"]
            y_bar2 = self.ncr_data["TOTAL_JOB_NCRS"]
            # get percentages
            y_line1 = [round((b / a) * 100, 1) for a, b in zip(y_bar1, y_bar2)]
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # pass into to 2nd chart
            left_y_label2 = "Count of Failure Categories"
            title2 = "Pareto of Top JOB Failure Categories\n" + current_monthyear
            bar_legend2 = [current_monthyear]

            # get categories and values for each
            categories, y_bar2 = self.get_graph_categories(self.total_jobs_ncr_data)
            # update graph
            self.update_lastmonth_graph(categories, y_bar2, left_y_label2, title2, bar_legend2)

            # update table
            self.update_table(self.total_jobs_ncr_data)

        if self.combo_chart.currentText() == "PO Line Items":
            left_y_label = "Count by Month"
            right_y_label = "% PO Line Items Non-Conforming"
            title = "% PO Line Items Impacted by NCRs by Month"
            bar_legend = ["PO LI Rec", "PO NCRs"]
            line_legend = ["% NCRs"]
            y_bar1 = self.qc_data["PO_LINE_ITEMS"]
            y_bar2 = self.ncr_data["TOTAL_PO_NCRS"]
            # get percentages
            y_line1 = [round((b / a) * 100, 1) for a, b in zip(y_bar1, y_bar2)]
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # pass into to 2nd chart
            left_y_label2 = "Count of Failure Categories"
            title2 = "Pareto of Top Purchased Part Failure Categories\n" + current_monthyear
            bar_legend2 = [current_monthyear]

            # get categories and values for each
            categories, y_bar2 = self.get_graph_categories(self.po_ncr_data)
            # update graph
            self.update_lastmonth_graph(categories, y_bar2, left_y_label2, title2, bar_legend2)

            # update table
            self.update_table(self.po_ncr_data)

        if self.combo_chart.currentText() == "Total PO Items":
            left_y_label = "Count by Month"
            right_y_label = ""
            title = "Qty of PO Items Received by Month"
            bar_legend = ["Qty"]
            line_legend = ["1"]
            y_bar1 = self.qc_data["TOTAL_RECEIVED"]
            y_bar2 = None
            y_line1 = None
            self.update_months_graph(left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1)

            # update table
           # self.update_table(self.po_ncr_data)

            # clear 2nd figure and table
            self.tableWidget.clear()
            self.figure_lastmonth.clear()


        # emit for changing the window size slightly due to graphs being changed
        # for some reason on initial draw of graph loadup the elements in the graph are off... but this fixes it....
        self.resize_window.emit("change size")

    # update the months graph
    def update_months_graph(self, left_y_label, right_y_label, title, bar_legend, line_legend, y_bar1, y_bar2, y_line1):

        # transform x labels to make them a little smaller
        new_months = []
        for i in self.chart_months:
            renamed_month = i.split()[1][0:3]
            new_months.append(i.split()[0] + " " + renamed_month)

        # clear figure and any plots in it
        self.figure_12_month.clear()

        # create axis and set font
        self.ax_12_month = self.figure_12_month.add_subplot(111)

        # plot bar graphs
        self.ax_12_month.bar(new_months, y_bar1, width=-0.4, align='edge', color='lightblue')
        self.ax_12_month.bar_label(self.ax_12_month.containers[0], label_type='center', rotation=90, fontsize=7)

        if y_bar2 is not None:
            self.ax_12_month.bar(new_months, y_bar2, width=0.4, align='edge', color='lightgreen')
            self.ax_12_month.bar_label(self.ax_12_month.containers[1], label_type='edge', fontsize=7)

        self.ax_12_month.yaxis.grid(True, linestyle='-', which='both', color='gray', alpha=0.2)

        # this needed for 1 to 1 mapping between tick positions and tick labels... or you start getting errors
        tick_positions = range(len(new_months))
        self.ax_12_month.set_xticks(tick_positions)

        # set the x tick labels for graph
        self.ax_12_month.set_xticklabels(new_months, rotation=25, ha='right')

        # set y label and title for graph
        self.ax_12_month.set_ylabel(left_y_label, fontweight="bold")
        self.ax_12_month.set_title(title, fontweight="bold", fontsize=12)

        # Set font size for x-axis tick labels
        self.ax_12_month.tick_params(axis='x', labelsize=6)

        # Set font size for y-axis tick labels
        self.ax_12_month.tick_params(axis='y', labelsize=7)

        if y_line1 is not None:
            # line graph, need to call twinx to put a line graph and bar graph together
            ax2 = self.ax_12_month.twinx()

            # get the line graph max value and add to it, to set the y limit for the graph slightly beyond the dataset
            # formula increases y-tick max by 50% more of wahtever my max value is
            y_max = max(y_line1) + (max(y_line1) * .5)

            # plot linegraph
            ax2.plot(new_months, y_line1, marker="*", color="orange")

            # check to make sure y_max won't be infinite
            if "inf" in str(y_max):
                y_max = 0

            # ensure y ticks start at 0 and are a little beyond my dataset
            ax2.set_ylim(0, y_max)

            # set right label axis of graph
            ax2.yaxis.set_label_position("right")
            ax2.yaxis.tick_right()
            ax2.set_ylabel(right_y_label, fontweight="bold", fontsize=7)

            # add % symbol to the yticks of the graph
            ax2.yaxis.set_major_formatter(mticker.FuncFormatter(self.percent_formatter))

            # Format and add labels above each data point in the linegraph
            for i, (xi, yi) in enumerate(zip(new_months, y_line1)):
                label = f'{yi:.1f}%'
                ax2.annotate(label, (xi, yi), textcoords="offset points", xytext=(5, 3), ha='center', fontsize=6)

            # Set font size for y-axis tick labels
            ax2.tick_params(axis='y', labelsize=7)

        # adjust graph position slightly in the UI to make room for legends at bottom
        box = self.ax_12_month.get_position()
        self.ax_12_month.set_position([box.x0, box.y0 + box.height * 0.13, box.width, box.height * 0.9])

        # Create a list of handles and labels for the legend
        handles, labels = [], []

        # Add bar legend entries
        for i, label in enumerate(bar_legend):
            handles.append(self.ax_12_month.containers[i])
            labels.append(label)

        if y_line1 is not None:
            # Add line legend entries
            for i, label in enumerate(line_legend):
                handles.append(ax2.lines[i])
                labels.append(label)

        # Create a combined legend for both axes
        self.ax_12_month.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, -0.18),
                                fancybox=True, shadow=True, ncol=3, fontsize=6)

        # refresh canvas
        self.canvas_12_month.draw()

    # function just for adding % sign to the y labels on the graphs
    def percent_formatter(self, y, pos):
        return f'{y:.1f}%'

    # update the lastmonth graph
    def update_lastmonth_graph(self, categories, y_bar1, left_y_label, title, bar_legend):
        # clear figure and any plots in it
        self.figure_lastmonth.clear()

        # create axis and set font
        self.ax_lastmonth = self.figure_lastmonth.add_subplot(111)

        # plot bar graphs
        self.ax_lastmonth.bar(categories, y_bar1, width=-0.6, align='center', color='orange')

        self.ax_lastmonth.bar_label(self.ax_lastmonth.containers[0], label_type='edge', fontsize=7)

        self.ax_lastmonth.yaxis.grid(True, linestyle='-', which='both', color='gray', alpha=0.2)

        # this needed for 1 to 1 mapping between tick positions and tick labels... or you start getting errors
        tick_positions = range(len(categories))
        self.ax_lastmonth.set_xticks(tick_positions)

        # set the x tick labels for graph
        self.ax_lastmonth.set_xticklabels(categories)

        # set y label and title for graph
        self.ax_lastmonth.set_ylabel(left_y_label, fontweight="bold")
        self.ax_lastmonth.set_title(title, fontweight="bold", fontsize=12)

        # Set font size for x-axis tick labels
        self.ax_lastmonth.tick_params(axis='x', labelsize=6)

        # Set font size for y-axis tick labels
        self.ax_lastmonth.tick_params(axis='y', labelsize=7)

        # calculation for upping max value on graph
        y_max = max(y_bar1) + (max(y_bar1) * .1)

        # check to make sure y_max won't be infinite
        if "inf" in str(y_max):
            y_max = 0

        # ensure y ticks start at 0 and are a little beyond my dataset
        self.ax_lastmonth.set_ylim(0, y_max)

        # adjust graph position slightly in the UI to make room for legends at bottom
        box = self.ax_lastmonth.get_position()
        self.ax_lastmonth.set_position([box.x0, box.y0 + box.height * 0.13, box.width, box.height * 0.9])

        # Create a list of handles and labels for the legend
        handles, labels = [], []

        # Add bar legend entries
        for i, label in enumerate(bar_legend):
            handles.append(self.ax_lastmonth.containers[i])
            labels.append(label)

        # Create a combined legend for both axes
        self.ax_lastmonth.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, -0.18),
                                fancybox=True, shadow=True, ncol=3, fontsize=6)

        # refresh canvas
        self.canvas_lastmonth.draw()

    # get top 6 categories and their values from dataframe and return the categories and counts
    def get_graph_categories(self, df):
        # filter out None values
        filtered = df[~df["Failure Category"].isin(["None"])]

        # get top 6 groups and their counts
        top_6 = filtered.groupby("Failure Category").size().nlargest(6).reset_index(name='count')

        # put data to lists for passing to charts
        categories = top_6["Failure Category"].to_list()
        count = top_6["count"].to_list()

        # run check on string length and modify string for graphs... dont' want them too long
        altered_categories = []
        for i in categories:
            if len(i) > 15:
                if " " in i:
                    values = i.split(" ", 1)
                    value = "\n".join(values)
                    altered_categories.append(value)
            else:
                altered_categories.append(i)

        return altered_categories, count

    # populate table widget
    def update_table(self, df):
        # update dataframe dates to remove timestamps for any columns with "date" in it
        date_columns = df.filter(like='Date')

        for column in date_columns:
            df[column] = df[column].astype(str).str.split().str[0]

       # print(df)
        # get how many rows / columns
        num_rows, num_cols = df.shape

        # set rows and columns
        self.tableWidget.setRowCount(num_rows)
        self.tableWidget.setColumnCount(num_cols)

        # font for items being added to qtablewidget
        item_font = QFont()
        item_font.setPointSize(7)

        # populate tablewidget
        for row in range(num_rows):
            for col in range(num_cols):
                value = str(df.iat[row, col])

                # change value if it's one of these
                if "NaT" in str(df.iat[row, col]) or "None" in str(df.iat[row, col]):
                    value = "TBD and/or NA"

                item = QTableWidgetItem(value)
                item.setFont(item_font)

                self.tableWidget.setItem(row, col, item)

        self.tableWidget.setWordWrap(False)

        # set column headers
        self.tableWidget.setHorizontalHeaderLabels(df.columns)

        # make column headers bold
        header_font = QFont()
        header_font.setBold(True)
        self.tableWidget.horizontalHeader().setFont(header_font)

        # resize cells to fit contents
        self.tableWidget.resizeRowsToContents()

    # pass charts and tables to temp excel to use as a medium for printing
    def setup_print_charts(self):
        # get current choice in combobox for which dataframe to pass to excel for printing
        current_charts = self.combo_chart.currentText()

        self.charts_available = ["PK JOBs", "AK JOBs", "FA JOBs", "SA JOBs", "FA SEQs", "Overall JOBs",
                                 "PO Line Items", "Total PO Items"]

        if "Total PO" not in current_charts:
            title_pref = ""
            if "PK" in current_charts:
                title_pref = "PK JOBs NCR Review "
            elif "AK" in current_charts:
                title_pref = "AK JOBs NCR Review "
            elif "FA JOBs" in current_charts:
                title_pref = "FA JOBs NCR Review "
            elif "SA JOBs" in current_charts:
                title_pref = "SA JOBs NCR Review "
            elif "FA SEQs" in current_charts:
                title_pref = "FA SEQs NCR Review "
            elif "Overall" in current_charts:
                title_pref = "Overall JOBs NCR Review "
            elif "PO Line" in current_charts:
                title_pref = "Purchase Order NCR Review "

            title = title_pref + self.chart_months[0].split()[1] + " - " + self.chart_months[-1].split()[1] + \
                    " " + self.chart_months[-1].split()[0]


            temp_path = "C:\\"
            # save current charts
            self.figure_12_month.savefig(temp_path + "12month.png", dpi=300)
            self.figure_lastmonth.savefig(temp_path + "lastmonth.png", dpi=300)

            # =================== get column headers based on visual ======================================
            # =================== get column headers based on visual ======================================
            # =================== get column headers based on visual ======================================
            # convert content of Qtablewidget to dataframe to then easily pass that to the excel file
            rowcount = self.tableWidget.rowCount()

            data = []

            # After the user has rearranged the columns, retrieve the updated order and updated header value order
            # header.logicalIndex must be used to do this based on the users visual changes if they move the columns
            # otherwise parsing the qtablewidget will return the original columns
            column_order = []
            header_values = []
            header = self.tableWidget.horizontalHeader()

            # Get the header index for the current visual index
            for visual_index in range(header.count()):
                logical_index = header.logicalIndex(visual_index)
                column_order.append(logical_index)

                # Get the header value for the current visual index
                header_item = self.tableWidget.horizontalHeaderItem(logical_index)
                header_value = header_item.text() if header_item else ""
                header_values.append(header_value)

            # iterate through table to get data
            for row in range(rowcount):
                rowData = []
                for column in column_order:
                    widgetItem = self.tableWidget.item(row, column)
                    if widgetItem and widgetItem.text:
                        rowData.append(widgetItem.text())
                    else:
                        rowData.append('NULL')

                data.append(rowData)
            # =================== get column headers based on visual ======================================
            # =================== get column headers based on visual ======================================
            # =================== get column headers based on visual ======================================

            # convert qtablewidget data list to dataframe for passing to excel
            table_dataframe = pd.DataFrame(data)

            # set column names of dataframe
            table_dataframe.columns = header_values

            # ================================calculating image resize================================
            # ================================calculating image resize================================
            # ================================calculating image resize================================
            # get what i want the column widths to be for the excel... pass to list , this will be used later
            # for scaling images based onthe width of the combined columns
            column_widths = []
            for i, col in enumerate(table_dataframe.columns):
                max_len = max(table_dataframe[col].astype(str).str.len())
                if len(col) > max_len:
                    column_width = len(col) + 2  # Add a little extra space
                else:
                    column_width = max_len + 2
                column_widths.append(column_width)

            # defaults for excel for determining pixel lengths
            excel_default_width = 64/8.43
            excel_default_height = 20

            # get pixel width of all columns
            total_pixels_width = int(excel_default_width*sum(column_widths))

            # pixel width i want for each graph
            image_width_iwant = int(total_pixels_width/2) - 32 - 5

            # get image size
            img_12month = Image.open(temp_path + "12month.png")
            img_width_12month, img_height_12month = img_12month.size

            # get ratio between what i want the image size to be and actual image size to scale height correctly
            width_ratio_12month = image_width_iwant / img_width_12month

            # new height that i want
            image_height_iwant = int(img_height_12month*width_ratio_12month)

            # not using the below code for reszing the images... as it produced an inferior
            # image quality when resizing, but keep code just in case
            """
            resized = img_12month.resize((image_width_iwant, image_height_iwant))
            resized.save(temp_path + "12month.png")

            # get 2nd image for recalculating size
            img_last = Image.open(temp_path + "lastmonth.png")
            img_width_last, img_height_last = img_last.size

            # get ratio between what i want the image size to be and actual image size to scale height correctly
            width_ratio_last = image_width_iwant / img_width_last

            # new height that i want
            image_height_iwant = int(img_height_last * width_ratio_last)

            resized = img_last.resize((image_width_iwant, image_height_iwant))
            resized.save(temp_path + "lastmonth.png")
            """
            # ================================calculating image resize================================
            # ================================calculating image resize================================
            # ================================calculating image resize================================

            # ==================== creation of excel =================================================
            # ==================== creation of excel =================================================
            # ==================== creation of excel =================================================
            # create excel file
            writer = pd.ExcelWriter(temp_path + "temp_excel_wr.xlsx", engine="xlsxwriter")

            # get which row to start passing dataframe to, based on height of recalculated images
            # (+2) represents counting 2 additional rows for space
            row_start = int((image_height_iwant / excel_default_height) + 2)

            # write table to excel
            table_dataframe.to_excel(writer, index=False, startrow=row_start+1, startcol=0)

            wb = writer.book
            ws = writer.sheets["Sheet1"]

            # Add borders to all cells in the table
            border_format = wb.add_format({'border': 1})  # 1 represents a thin border
            ws.conditional_format(row_start+1, 0, len(table_dataframe)+row_start+1, len(table_dataframe.columns), {'type': 'no_blanks', 'format': border_format})

            # Iterate through the columns and set the column width to fit the content
            for i in range(len(column_widths)):
                ws.set_column(i, i, column_widths[i])

            # header formating
            header_format = wb.add_format({"bold": True, "fg_color": "#D7E4BC", "border": 1, })

            # change the formatting of the header items to look better in the excel file
            for col_num, value in enumerate(table_dataframe.columns.values):
                ws.write(row_start+1, col_num, value, header_format)

            # format for adding title to excel
            bold_format = wb.add_format({'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_size': 26})

            # merge cells
            ws.merge_range(0, 0, 0, len(column_widths), title, bold_format)

            # offset for 2nd image
            offset = image_width_iwant + 5 + 50

            # inserting image using openpyxl to produce superior image quality, but keeping code just in case
            # should have just used openpyxl instead of xlsxwriter
            """
            # insert images
            ws.insert_image('B2', temp_path + "12month.png")
            ws.insert_image('B2', temp_path + "lastmonth.png", {"x_offset": offset})
            """

            # do page setup to set margins and set to landscape
            ws.set_margins(.5, 0, 0, 0)
            ws.set_landscape()

            # set to page on width only
            ws.fit_to_pages(1, 0)

            writer.close()

            # the below code uses openpyxl to insert the images... rather than using xlsxwriter
            # this is the only way i could get a better/clearer image... i tried other options with resizing
            # and inserting image... but this method produces clearest picture
            # ..... i should have just used openpyxl instead of xlsxwriter... oh well

            wb = load_workbook(temp_path + "temp_excel_wr.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")

            # set minimum width/height variables if user chooses only a few ncr columns
            if image_width_iwant < 250:
                image_width_iwant = 250
            if image_height_iwant < 125:
                image_width_iwant = 125


            # set width and hieght of images
            logo = Im(temp_path + "12month.png")
            logo.height = image_height_iwant
            logo.width = image_width_iwant + 50

            # add image
            ws.add_image(logo, "A2")

            logo = Im(temp_path + "lastmonth.png")
            logo.height = image_height_iwant
            logo.width = image_width_iwant + 50

            # i don't exactly knoow how the below code works with absoluteanchor... but it works
            # any other method with anchoring causes a crash
            p2e = pixels_to_EMU
            position = XDRPoint2D(p2e(offset), p2e(45))
            size = XDRPositiveSize2D(p2e(logo.width), p2e(logo.height))
            logo.anchor = AbsoluteAnchor(pos=position, ext=size)

            ws.add_image(logo)

            wb.save(temp_path + "temp_excel_wr.xlsx")

            wb.close()

        # if total po, i was just the 1 chart saved, nothing else, no tables nor 2nd chart
        if "Total PO" in current_charts:
            temp_path = "C:\\"
            # save current charts

            self.figure_12_month.savefig(temp_path + "12month.png", dpi=300)

            # create excel file
            writer = pd.ExcelWriter(temp_path + "temp_excel_wr.xlsx", engine="xlsxwriter")

            wb = writer.book
            wb.add_worksheet()
            ws = writer.sheets["Sheet1"]

            # do page setup to set margins and set to landscape
            ws.set_margins(0, 0, 0, 0)
            ws.set_landscape()

            # set to page on width only
            ws.fit_to_pages(1, 0)

            writer.close()

            # below to use openpyxl instead of xlsxwriter for reszing and inserting image... better image quality
            wb = load_workbook(temp_path + "temp_excel_wr.xlsx")
            ws = wb.get_sheet_by_name("Sheet1")

            logo = Im(temp_path + "12month.png")
            logo.height = 720
            logo.width = 1080

            p2e = pixels_to_EMU

            position = XDRPoint2D(p2e(64), p2e(20))
            size = XDRPositiveSize2D(p2e(logo.width), p2e(logo.height))

            logo.anchor = AbsoluteAnchor(pos=position, ext=size)

            ws.add_image(logo)

            wb.save(temp_path + "temp_excel_wr.xlsx")

            wb.close()


        # emit to UI to get printer selection, this will then call a UI function that activates print_charts function here
        self.print_selection.emit("print")

    # actually print charts
    def print_charts(self, printer):
        # get default printer, set back to this after printing
        default_printer = win32print.GetDefaultPrinter()

        #path to excel file that i want to print
        excel_path = "C:\\temp_excel_wr.xlsx"

        # print excel FAI, set all sheets to fit to page, reset default printer back to original
        if "None" not in printer:
            # change to printer selected
            win32print.SetDefaultPrinter(printer)

            # use Excel API to set all pages to fit to page before printing
            o = win32com.client.Dispatch('Excel.Application')
            o.visible = False
            wb = o.Workbooks.Open(excel_path)

            # print
            wb.PrintOut()
            wb.Close()

            # set back to default printer
            win32print.SetDefaultPrinter(default_printer)
        else:
            # msg if no printer selected in UI
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Pick a Printer first!")
            msg.setWindowTitle("Select Printer")
            msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
            msg.exec()


# =============================chart page===========================================================
# =============================chart page===========================================================
# =============================chart page===========================================================


# =============================add month page========================================================
# =============================add month page========================================================
# =============================add month page========================================================
    # add month page on UI
    def add_month_data_page(self):
        self.month_scroll = QScrollArea()
        self.month_scroll.setWidgetResizable(True)
        self.month_widget = QWidget()

        self.verticalLayout_month = QVBoxLayout()

        myfont = QFont()
        myfont.setPointSize(12)

        self.label_month = QLabel("<b>Add Month Of Data</b>")
        self.label_month.setFont(myfont)
        self.label_month.setAlignment(Qt.AlignHCenter)

        self.verticalLayout_month.addWidget(self.label_month)
        self.verticalLayout_month.addSpacing(50)

        self.lineEdit_job = QLineEdit()
        self.lineEdit_job.setPlaceholderText("Enter Path to JOB Data from EpicQuality as xls, xlsx or csv file......")

        self.verticalLayout_month.addWidget(self.lineEdit_job)
        self.verticalLayout_month.addSpacing(20)

        self.lineEdit_ncr = QLineEdit()
        self.lineEdit_ncr.setPlaceholderText("Enter Path to NCR Data from EpicQuality or Enquiries as xls, xlsx or csv file......")

        self.verticalLayout_month.addWidget(self.lineEdit_ncr)
        self.verticalLayout_month.addSpacing(20)

        self.lineEdit_po = QLineEdit()
        self.lineEdit_po.setPlaceholderText("Enter Path to QC REC Data from EpicQueries as xls, xlsx or csv file......")

        self.verticalLayout_month.addWidget(self.lineEdit_po)
        self.verticalLayout_month.addSpacing(20)

        self.pushButton_month = QPushButton("Add Month")
        self.pushButton_month.setToolTip("NOTE: If adding older data, you will need to reload program!\n"
                                         "If adding new data or replacing existing data, no need to reload.")
        self.pushButton_month.clicked.connect(self.collect_data)

        self.verticalLayout_month.addWidget(self.pushButton_month, alignment=Qt.AlignHCenter, stretch=0)

        self.label_2_month = QLabel("")
        self.label_2_month.setAlignment(Qt.AlignHCenter)

        self.verticalLayout_month.addWidget(self.label_2_month)
        self.verticalLayout_month.addStretch()

        self.month_widget.setLayout(self.verticalLayout_month)

        self.month_scroll.setWidget(self.month_widget)

        self.addWidget(self.month_scroll)

    # collect data from excels by first checking excel extensions and all files entered in
    def collect_data(self):
        available_ext = ["xlsx", "xls", "csv"]

        if len(self.lineEdit_job.text()) != 0 and len(self.lineEdit_po.text()) != 0 and len(self.lineEdit_ncr.text()) != 0:
            # remove any quotes from doing a shift+copy path in windows
            job_path = self.lineEdit_job.text().replace('"', "")
            ncr_path = self.lineEdit_ncr.text().replace('"', "")
            po_path = self.lineEdit_po.text().replace('"', "")

            # split to get extension at end
            job_split = job_path.split(".")
            ncr_split = ncr_path.split(".")
            po_split = po_path.split(".")

            if job_split[-1] not in available_ext or ncr_split[-1] not in available_ext or po_split[-1] not in available_ext:

                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Detected Invalid extension type!")
                msg.setWindowTitle("Error")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec()
            else:
                self.add_month_database(job_path, ncr_path, po_path)
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Must have all data before adding months data!")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec()

    # get data to ncr database
    def add_month_database(self, job_path, ncr_path, po_path):
        # make sure path to files exists, then check to make sure each file contains
        # the necessary columns before adding to database
        if os.path.exists(job_path) and os.path.exists(ncr_path) and os.path.exists(po_path):

            # get table columns from UI and ncr database for comparison against excels
            job_tables, ncr_tables, qc_tables = self.get_tables_comparison()

            # get extension
            job_split = job_path.split(".")
            ncr_split = ncr_path.split(".")
            po_split = po_path.split(".")

            # put what i want to iterate into a list of lists for easier iteration
            # note:  SIGNEDOFFDATETIME column for JOBS
            #        Date column for NCRs
            #        DAT_ column for qc rec data
            month_data_iterate = [[" JOB", job_path, job_split, "SIGNEDOFFDATETIME", job_tables],
                                  [" NCR", ncr_path, ncr_split, "Date", ncr_tables],
                                  [" QC", po_path, po_split, "DAT_", qc_tables]]

            # var that increments when columns are checked
            test_continue = 0

            # append year_month to list this will be used as a check to make sure all year/months between files match
            year_month_list = []

            # put dataframes in a list for iteration later
            dataframes_months = []

            # iterate list
            for i in month_data_iterate:
                # get dataframes based on extension
                if "csv" in i[2]:
                    df = pd.read_csv(i[1], engine="pyarrow")
                    dataframes_months.append(df)
                    year_month = datetime.strptime(str(df[i[3]][0]).split(" ")[0], "%m/%d/%Y")
                    table_name = str(year_month.year) + " " + str(year_month.month)
                    year_month_list.append(table_name)

                if "xls" in i[2] or "xlsx" in i[2]:
                    df = pd.read_excel(i[1])
                    dataframes_months.append(df)
                    year_month = datetime.strptime(str(df[i[3]][0]).split(" ")[0], "%Y-%m-%d")
                    table_name = str(year_month.year) + " " + str(year_month.month)
                    year_month_list.append(table_name)

                # do comparison tests of columns to make sure what is it excel matches database tables
                if " JOB" in i[0] or " QC" in i[0]:
                    columns = df.columns.to_list()
                    if set(i[4]) == set(columns):
                        test_continue +=1
                if " NCR" in i[0]:
                    columns = df.columns.to_list()

                    # removing drawing no column from the column check, to provide compatability between exporting the data
                    # from epicenquiries or epic quality
                    i[4].remove("Drawing No")

                    # check to make sure all columns that can be chosen in UI are in the dataframe column list
                    if set(i[4]) <= set(columns):
                        test_continue +=1

            # check to make sure all files are the same month/year
            same_month_year = all(element == year_month_list[0] for element in year_month_list)

            # make sure all dataframe columns in sql table columns and files all have same month/year
            if test_continue == 3 and same_month_year == True:
                # create sql tables
                self.create_sql_tables(year_month_list[0], dataframes_months)

                # update combobox in UI with new month option
                self.update_combobox_options(year_month_list[0])

                # messaged when month o fdata successfully added
                msg = QMessageBox()
                msg.setText("Month of Data Added!")
                msg.setWindowTitle("Done")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec()

            else:
                # error if not 3 or not true
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Warning)
                msg.setText("Data In Files not correct!\n Either columns don't match to database\n or files have different months/year!")
                msg.setWindowTitle("Error")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec()

        # error message if path to 1 of the files does not exists
        else:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Warning)
            msg.setText("Path to 1 of the files does not exist!")
            msg.setWindowTitle("Error")
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec()

    # get tables for JOBS/NCR/QC
    def get_tables_comparison(self):
        conn = sqlite3.connect("your database")
        cursor = conn.cursor()

        # get ncr columns to check from the UI, needs to be from the UI due to enquiries/epicquality differences
        # rest will be from the tables in the database

        ncr_columns = [self.list_ncr_data.item(i).text() for i in range(self.list_ncr_data.count())]

        # get last index in combobox from UI, can use this to get the table names to query from database
        combo_value = self.comboBox_start.itemText(self.comboBox_start.count()-1)
        split_combo_Value = combo_value.split()

        # get the month as an integer
        index = [x for x, item in enumerate(calendar.month_name) if item == split_combo_Value[1]]

        # table names to query in database
        job_table_name = "'" + split_combo_Value[0] + " " + str(index[0]) + " " + "JOB'"
        vendor_table_name = "'" + split_combo_Value[0] + " " + str(index[0]) + " " + "QC'"

        job_query = "PRAGMA table_info(" + job_table_name + ");"

        cursor.execute(job_query)

        # get job table names
        fetch = cursor.fetchall()

        # pass table names as a list minus the "index" columns name
        job_columns = [row[1] for row in fetch][1:]

        vendor_query = "PRAGMA table_info(" + vendor_table_name + ");"

        cursor.execute(vendor_query)

        # get all table names
        fetch = cursor.fetchall()

        conn.close()

        # pass table names as a list minus the "index" columns name
        vendor_columns = [row[1] for row in fetch][1:]

        return job_columns, ncr_columns, vendor_columns

    # send dataframe tables to sql table
    def create_sql_tables(self, year_month, dataframes):
        conn = sqlite3.connect("your database")
        cursor = conn.cursor()

        # create tables on the sql database
        job_table_name = year_month + " JOB"
        dataframes[0].to_sql(job_table_name, conn, if_exists='replace')

        ncr_table_name = year_month + " PO"
        dataframes[1].to_sql(ncr_table_name, conn, if_exists='replace')

        qc_table_name = year_month + " QC"
        dataframes[2].to_sql(qc_table_name, conn, if_exists='replace')

        # create index on job table
        job_index = "index " + job_table_name
        query = "CREATE INDEX '" + job_index + "' ON '" + job_table_name + "' (JOBNO, REFID, WORKCENTERNAME, PROCESS_ID);"

        cursor.execute(query)

        conn.commit()
        conn.close()

    # update combobox year/month options
    def update_combobox_options(self, year_month):
        # reconfigure month/year to pass to combobox
        month = calendar.month_name[int(year_month.split()[1])]
        pass_to_combo = year_month.split()[0] + " " + month

        # check to see if item already exists in combobox (if replacing a month)
        index = self.comboBox_start.findText(pass_to_combo)

        if index != -1:
            pass
        else:
            # add to the comboboxes in UI with new month option
            self.comboBox_start.addItem(pass_to_combo)
            self.comboBox_end.addItem(pass_to_combo)

# =============================add month page========================================================
# =============================add month page========================================================
# =============================add month page========================================================


# =============================filter page========================================================
# =============================filter page========================================================
# =============================filter page========================================================
    # filter page on UI
    def data_filter_page(self):
        self.filter_scroll = QScrollArea()
        self.filter_scroll.setWidgetResizable(True)

        self.filter_widget = QWidget()
        self.verticalLayout_4 = QVBoxLayout()

        self.label_4 = QLabel("<b>Select JOB Filter Options</b>")
        self.label_4.setAlignment(Qt.AlignHCenter)

        myfont = QFont()
        myfont.setPointSize(12)
        self.label_4.setFont(myfont)

        self.verticalLayout_4.addWidget(self.label_4)

        self.groupBox = QGroupBox()

        self.horizontalLayout = QHBoxLayout()
        self.horizontalLayout.setContentsMargins(20, 30, 20, 30)

        self.radioButton_tom_filter = QRadioButton("Tom H. Filters (default)")
        self.radioButton_tom_filter.setChecked(True)
        self.radioButton_tom_filter.clicked.connect(self.filter_list_updates)
        self.radioButton_filter = QRadioButton("Do Custom Filtering")
        self.radioButton_filter.clicked.connect(self.filter_list_updates)

        self.horizontalLayout.addStretch()
        self.horizontalLayout.addWidget(self.radioButton_tom_filter)
        self.horizontalLayout.addWidget(self.radioButton_filter)
        self.horizontalLayout.addStretch()

        self.groupBox.setLayout(self.horizontalLayout)

        self.verticalLayout_4.addWidget(self.groupBox)
        self.verticalLayout_4.addSpacing(15)

        self.label_5 = QLabel("<b>JOB Filter Options</b>       (for custom filtering)")
        self.label_5.setAlignment(Qt.AlignHCenter)

        myfont = QFont()
        myfont.setPointSize(10)
        self.label_5.setFont(myfont)

        self.verticalLayout_4.addWidget(self.label_5)

        self.line = QFrame()
        self.line.setFrameShape(QFrame.HLine)
        self.line.setFrameShadow(QFrame.Sunken)

        self.verticalLayout_4.addWidget(self.line)
        self.verticalLayout_4.addSpacing(15)

        self.horizontalLayout_5 = QHBoxLayout()
        self.verticalLayout = QVBoxLayout()

        self.job_vertical_layout = QVBoxLayout()
        self.job_options = QLabel("<b>JOB Options</b>")
        self.job_options.setAlignment(Qt.AlignHCenter)

        self.job_vertical_layout.addWidget(self.job_options)

        self.job_groupbox = QGroupBox()

        self.job_group_layout = QVBoxLayout()

        self.checkBox_fa = QCheckBox("Count Total FA SEQs \nBefore Filters are Applied")
        self.checkBox_fa.setCheckState(Qt.Unchecked)
        self.checkBox_fa.setEnabled(False)

        self.job_group_layout.addSpacing(50)
        self.job_group_layout.addWidget(self.checkBox_fa)
        self.job_group_layout.addStretch()

        self.job_groupbox.setLayout(self.job_group_layout)

        self.job_vertical_layout.addWidget(self.job_options)
        self.job_vertical_layout.addWidget(self.job_groupbox)
        self.job_vertical_layout.addStretch()
        self.horizontalLayout_5.addLayout(self.job_vertical_layout)

        self.label = QLabel("<b>Work Centers</b>")
        self.label.setAlignment(Qt.AlignHCenter)

        self.verticalLayout.addWidget(self.label)

        self.list_workcenter = QListWidget()

        self.verticalLayout.addWidget(self.list_workcenter)

        self.horizontalLayout_2 = QHBoxLayout()

        self.pushButton = QPushButton("Select All")
        self.pushButton.setObjectName("Select Work")
        self.pushButton.clicked.connect(self.filter_buttons)
        self.pushButton.setEnabled(False)
        self.pushButton_2 = QPushButton("Clear")
        self.pushButton_2.setEnabled(False)
        self.pushButton_2.setObjectName("Clear Work")
        self.pushButton_2.clicked.connect(self.filter_buttons)

        self.horizontalLayout_2.addWidget(self.pushButton)
        self.horizontalLayout_2.addWidget(self.pushButton_2)

        self.verticalLayout.addLayout(self.horizontalLayout_2)

        self.horizontalLayout_5.addLayout(self.verticalLayout)

        self.verticalLayout_2 = QVBoxLayout()

        self.label_2 = QLabel("<b>Process ID's</b>")
        self.label_2.setAlignment(Qt.AlignHCenter)

        self.verticalLayout_2.addWidget(self.label_2)

        self.list_process = QListWidget()

        self.verticalLayout_2.addWidget(self.list_process)

        self.horizontalLayout_3 = QHBoxLayout()

        self.pushButton_3 = QPushButton("Select All")
        self.pushButton_3.setEnabled(False)
        self.pushButton_3.setObjectName("Select Process")
        self.pushButton_3.clicked.connect(self.filter_buttons)
        self.pushButton_4 = QPushButton("Clear")
        self.pushButton_4.setEnabled(False)
        self.pushButton_4.setObjectName("Clear Process")
        self.pushButton_4.clicked.connect(self.filter_buttons)

        self.horizontalLayout_3.addWidget(self.pushButton_3)
        self.horizontalLayout_3.addWidget(self.pushButton_4)

        self.verticalLayout_2.addLayout(self.horizontalLayout_3)

        self.horizontalLayout_5.addLayout(self.verticalLayout_2)

        self.verticalLayout_3 = QVBoxLayout()
        self.label_3 = QLabel("<b>JOB Types</b>")
        self.label_3.setAlignment(Qt.AlignHCenter)

        self.verticalLayout_3.addWidget(self.label_3)

        self.list_type = QListWidget()

        self.verticalLayout_3.addWidget(self.list_type)

        self.horizontalLayout_4 = QHBoxLayout()

        self.pushButton_5 = QPushButton("Select All")
        self.pushButton_5.setEnabled(False)
        self.pushButton_5.setObjectName("Select Job")
        self.pushButton_5.clicked.connect(self.filter_buttons)
        self.pushButton_6 = QPushButton("Clear")
        self.pushButton_6.setEnabled(False)
        self.pushButton_6.setObjectName("Clear Job")
        self.pushButton_6.clicked.connect(self.filter_buttons)

        self.horizontalLayout_4.addWidget(self.pushButton_5)
        self.horizontalLayout_4.addWidget(self.pushButton_6)

        self.verticalLayout_3.addLayout(self.horizontalLayout_4)
        self.horizontalLayout_5.addLayout(self.verticalLayout_3)
        self.verticalLayout_4.addLayout(self.horizontalLayout_5)
        self.filter_widget.setLayout(self.verticalLayout_4)
        self.filter_scroll.setWidget(self.filter_widget)

        # add widget(Page) to Qstackedlayout
        self.addWidget(self.filter_scroll)

        # activate populating Qlistwidgets
        self.get_filter_data()

    # get filter data for data_filter_age from database
    def get_filter_data(self):
        # add sqlite3 connection to database for program to get data
        conn = sqlite3.connect("your database")
        cursor = conn.cursor()

        # get all tables with JOB in title and get last one for query
        query = "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%JOB' ORDER BY name"
        cursor.execute(query)

        # get last job table made
        last_job_table = "".join(cursor.fetchall()[-1])

        # close sql connection
        conn.close()

        # get distinct values from columns
        work_query = "SELECT DISTINCT workcentername FROM '" + last_job_table + "'"
        process_query = "SELECT DISTINCT process_id FROM '" + last_job_table + "'"
        job_query = "SELECT DISTINCT substr(jobno,1,4) AS jobs FROM '" + last_job_table + "'"

        # get data frames and sort for UI
        work_df = cx.read_sql(connect, work_query)
        work_df.sort_values(by=["WORKCENTERNAME"], inplace=True)
        process_df = cx.read_sql(connect, process_query)
        process_df.sort_values(by=["PROCESS_ID"], inplace=True)
        job_df = cx.read_sql(connect, job_query)
        job_df.sort_values(by=["jobs"], inplace=True)

        # get unique first 4 digits from jobno column, remove any digits
        job_types = job_df['jobs'].str.replace('\d+', '', regex=True).drop_duplicates(keep='first')

        # add items to list_workcenter
        for i in range(len(work_df)):
            item_value = work_df.iloc[i].values[0]
            item = QListWidgetItem(item_value)

            # start with items disabled
            item.setFlags(Qt.NoItemFlags)

            if any(x in item_value.upper() for x in workcenter_removal):
                item.setCheckState(Qt.Unchecked)
            else:
                item.setCheckState(Qt.Checked)

            self.list_workcenter.addItem(item)

        # add items to list_processid
        for i in range(len(process_df)):
            item_value = process_df.iloc[i].values[0]
            item = QListWidgetItem(item_value)

            # start with items disabled
            item.setFlags(Qt.NoItemFlags)

            if any(x in item_value.upper() for x in process_id_removal):
                item.setCheckState(Qt.Unchecked)
            else:
                item.setCheckState(Qt.Checked)

            self.list_process.addItem(item)

        # add items to list_jobtype
        for i in job_types:
            item = QListWidgetItem(i)
          #  item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)

            # start with items disabled
            item.setFlags(Qt.NoItemFlags)

            if any(x in i.upper() for x in jobnumber_test):
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)

            self.list_type.addItem(item)

    # update filter lists based on radio button hit
    def filter_list_updates(self):
        sender = self.sender()

        # update list states based on which radiobutton hit
        if "Custom" in sender.text():

            # disabling checking marking them all for now.. deciding i don't actually want to do this

            for index in range(self.list_workcenter.count()):
              #  self.list_workcenter.item(index).setCheckState(Qt.Checked)
                self.list_workcenter.item(index).setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            for index in range(self.list_process.count()):
              #  self.list_process.item(index).setCheckState(Qt.Checked)
                self.list_process.item(index).setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            for index in range(self.list_type.count()):
              #  self.list_type.item(index).setCheckState(Qt.Checked)
                self.list_type.item(index).setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)

            # enable buttons for lists
            self.pushButton.setEnabled(True)
            self.pushButton_2.setEnabled(True)
            self.pushButton_3.setEnabled(True)
            self.pushButton_4.setEnabled(True)
            self.pushButton_5.setEnabled(True)
            self.pushButton_6.setEnabled(True)
            self.checkBox_fa.setEnabled(True)

        # update list states based on which radiobutton hit
        if "Tom" in sender.text():
            for index in range(self.list_workcenter.count()):
                # start with items disabled
                self.list_workcenter.item(index).setFlags(Qt.NoItemFlags)

                if any(x in self.list_workcenter.item(index).text().upper() for x in workcenter_removal):
                    self.list_workcenter.item(index).setCheckState(Qt.Unchecked)
                else:
                    self.list_workcenter.item(index).setCheckState(Qt.Checked)

            for index in range(self.list_process.count()):
                # start with items disabled
                self.list_process.item(index).setFlags(Qt.NoItemFlags)

                if any(x in self.list_process.item(index).text().upper() for x in process_id_removal):
                    self.list_process.item(index).setCheckState(Qt.Unchecked)
                else:
                    self.list_process.item(index).setCheckState(Qt.Checked)

            for index in range(self.list_type.count()):
                # start with items disabled
                self.list_type.item(index).setFlags(Qt.NoItemFlags)

                if any(x in self.list_type.item(index).text().upper() for x in jobnumber_test):
                    self.list_type.item(index).setCheckState(Qt.Checked)
                else:
                    self.list_type.item(index).setCheckState(Qt.Unchecked)

            # enable buttons for lists
            self.pushButton.setEnabled(False)
            self.pushButton_2.setEnabled(False)
            self.pushButton_3.setEnabled(False)
            self.pushButton_4.setEnabled(False)
            self.pushButton_5.setEnabled(False)
            self.pushButton_6.setEnabled(False)
            self.checkBox_fa.setCheckState(Qt.Unchecked)
            self.checkBox_fa.setEnabled(False)

                    # for setup page of program, adding as i want every program to have a setup page

    # filter select/clear buttons for lists
    def filter_buttons(self):
        sender = self.sender()

        if "Select Work" in sender.objectName():
            for index in range(self.list_workcenter.count()):
                self.list_workcenter.item(index).setCheckState(Qt.Checked)
        if "Select Process" in sender.objectName():
            for index in range(self.list_process.count()):
                self.list_process.item(index).setCheckState(Qt.Checked)
        if "Select Job" in sender.objectName():
            for index in range(self.list_type.count()):
                self.list_type.item(index).setCheckState(Qt.Checked)

        if "Clear Work" in sender.objectName():
            for index in range(self.list_workcenter.count()):
                self.list_workcenter.item(index).setCheckState(Qt.Unchecked)
        if "Clear Process" in sender.objectName():
            for index in range(self.list_process.count()):
                self.list_process.item(index).setCheckState(Qt.Unchecked)
        if "Clear Job" in sender.objectName():
            for index in range(self.list_type.count()):
                self.list_type.item(index).setCheckState(Qt.Unchecked)
# =============================filter page========================================================
# =============================filter page========================================================
# =============================filter page========================================================


# =============================ncr page========================================================
# =============================ncr page========================================================
# =============================ncr page========================================================
    # ncr filter page
    def options_page(self):
        self.options_scroll = QScrollArea()
        self.options_scroll.setWidgetResizable(True)
        self.options_widget = QWidget()

        myfont = QFont()
        myfont.setPointSize(12)

        self.verticalLayout_3 = QVBoxLayout()

        self.label = QLabel("<b>Select NCR Filter Options<b>")
        self.label.setAlignment(Qt.AlignHCenter)
        self.label.setFont(myfont)

        myfont = QFont()
        myfont.setPointSize(10)

        self.verticalLayout_3.addWidget(self.label)

        self.groupBox = QGroupBox()

        self.horizontalLayout = QHBoxLayout()
        self.horizontalLayout.setContentsMargins(20, 30, 20, 30)

        self.radioButton_tom = QRadioButton("Tom H. Filters (default)")
        self.radioButton_tom.clicked.connect(self.change_options)
        self.radioButton_tom.setChecked(True)
        self.radioButton_custom = QRadioButton("Do Custom Filtering")
        self.radioButton_custom.clicked.connect(self.change_options)

        self.horizontalLayout.addStretch()
        self.horizontalLayout.addWidget(self.radioButton_tom)
        self.horizontalLayout.addWidget(self.radioButton_custom)
        self.horizontalLayout.addStretch()

        self.groupBox.setLayout(self.horizontalLayout)

        self.verticalLayout_3.addWidget(self.groupBox)
        self.verticalLayout_3.addSpacing(15)

        self.label_2 = QLabel("<b>NCR Filter Options</b> (for custom filtering)")
        self.label_2.setAlignment(Qt.AlignHCenter)
        self.label_2.setFont(myfont)
        self.verticalLayout_3.addWidget(self.label_2)

        self.line = QFrame()
        self.line.setFrameShape(QFrame.HLine)
        self.line.setFrameShadow(QFrame.Sunken)

        self.verticalLayout_3.addWidget(self.line)
        self.verticalLayout_3.addSpacing(20)

        self.horizontalLayout_2 = QHBoxLayout()

        self.verticalLayout_ncr_options = QVBoxLayout()

        myfont = QFont()
        myfont.setPointSize(8)

        self.ncr_options_label = QLabel("<b>Additional NCR Options</b>")
        self.ncr_options_label.setAlignment(Qt.AlignHCenter)
        self.ncr_options_label.setFont(myfont)

        self.groupBox_2 = QGroupBox()

        self.verticalLayout = QVBoxLayout()

        self.checkBox_drawing = QCheckBox("Use NCR Part# Field instead of Drawing# Field")
        self.checkBox_drawing.setToolTip("For When Sourcing Whether NCR is against a PK/AK/HK if fields are different")
        self.checkBox_linked = QCheckBox("PK99s on JOBs Count Towards Supplier NCRs")
        self.checkBox_linked.setToolTip("This will make total NCRs higher on the graphs, instead of just counting unique NCRs")
        self.checkBox_vendor = QCheckBox("Only Count NCRs that have a drawing \n(IE not against a vendor part#)")
        self.checkBox_vendor.setToolTip("Don't count NCRs against vendor part numbers")

        # set initial state of checkboxes
        self.checkBox_linked.setCheckState(Qt.Checked)
        self.checkBox_drawing.setEnabled(False)
        self.checkBox_linked.setEnabled(False)
        self.checkBox_vendor.setEnabled(False)

        self.verticalLayout.addSpacing(25)
        self.verticalLayout.addWidget(self.checkBox_drawing)
        self.verticalLayout.addSpacing(10)
        self.verticalLayout.addWidget(self.checkBox_linked)
        self.verticalLayout.addSpacing(10)
        self.verticalLayout.addWidget(self.checkBox_vendor)
        self.verticalLayout.addSpacing(10)
        self.verticalLayout.addStretch()
        self.verticalLayout.setAlignment(Qt.AlignHCenter)

        self.groupBox_2.setLayout(self.verticalLayout)

        self.verticalLayout_ncr_options.addWidget(self.ncr_options_label)
        self.verticalLayout_ncr_options.addWidget(self.groupBox_2)

        self.horizontalLayout_2.addSpacing(50)
        self.horizontalLayout_2.addLayout(self.verticalLayout_ncr_options, 50)
        self.horizontalLayout_2.addSpacing(50)

        self.verticalLayout_2 = QVBoxLayout()

        self.label_3 = QLabel("<b>NCR Data to collect</b>")
        self.label_3.setFont(myfont)
        self.label_3.setAlignment(Qt.AlignHCenter)

        self.verticalLayout_2.addWidget(self.label_3)

        self.list_ncr_data = QListWidget()

        self.verticalLayout_2.addWidget(self.list_ncr_data)

        self.horizontalLayout_2.addLayout(self.verticalLayout_2, 50)
        self.horizontalLayout_2.addSpacing(50)

        self.verticalLayout_3.addLayout(self.horizontalLayout_2)

        self.options_widget.setLayout(self.verticalLayout_3)

        self.options_scroll.setWidget(self.options_widget)

        self.addWidget(self.options_scroll)

        self.get_ncr_options()

    # get ncr options from database for qlist
    def get_ncr_options(self):

        # add sqlite3 connection to database for program to get data
        conn = sqlite3.connect("your database")
        cursor = conn.cursor()

        # get all tables with JOB in title and get last one for query
        query = "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%PO' ORDER BY name"
        cursor.execute(query)

        # get last ncr table made
        last_ncr_table = "".join(cursor.fetchall()[-1])

        # close sql connection
        conn.close()

        start = time.time()
        # get distinct values from columns
        ncr_query = "SELECT * FROM '" + last_ncr_table + "'"

        # get data frame
        ncr_df = cx.read_sql(connect, ncr_query)

        # get column names
        ncr_headers = list(ncr_df)

        # these are columns i want removed from being able to choose in UI.. this is so that
        # the user can export ncr data from enquiries or quality
        remove_columns = ['Linked Job/PO', 'Linked Part Number', 'Linked Disposition', 'Additional', 'SEQ', 'DISPBY',
         'REVIEW', 'ENGREVIEW', 'ASSIGNNOTE', 'DISPASSIGNNOTE', 'NCRPROGRESS', 'PRIORITY', 'FAAAPPROVE', 'FINALED',
         'SUBMITTED', 'SUBMITDISP', 'DBASUBMIT', 'Trend Review', 'QA Admin', 'Cause', 'Cause Info']

        # remove items from ncr_headers for UI, note the [1:] to remove the index column
        filtered_headers = [x for x in ncr_headers if x not in remove_columns][1:]
        filtered_headers.sort()

        # add items to list_ncr
        for i in range(len(filtered_headers)):
            item_value = filtered_headers[i]
            item = QListWidgetItem(item_value)

            # start with items disabled
            item.setFlags(Qt.NoItemFlags)

            if item_value in ncr_columns_display:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)

            self.list_ncr_data.addItem(item)

    # when switching between custom or Toms options
    def change_options(self):
        # items that can't be unchecked or program will crash as these items must be parsed in DATA class
        do_not_uncheck = ["NCR No", "Job No/PO", "Failure Category", "Drawing No"]

        sender = self.sender()

        # update list states based on which radiobutton hit
        if "Custom" in sender.text():
            for index in range(self.list_ncr_data.count()):
                if self.list_ncr_data.item(index).text() not in do_not_uncheck:
                    self.list_ncr_data.item(index).setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)

                self.checkBox_drawing.setEnabled(True)
                self.checkBox_linked.setEnabled(True)
                self.checkBox_vendor.setEnabled(True)

        # update list states based on which radiobutton hit
        if "Tom" in sender.text():
            for index in range(self.list_ncr_data.count()):
                # start with items disabled
                self.list_ncr_data.item(index).setFlags(Qt.NoItemFlags)

                if self.list_ncr_data.item(index).text() in ncr_columns_display:
                    self.list_ncr_data.item(index).setCheckState(Qt.Checked)
                else:
                    self.list_ncr_data.item(index).setCheckState(Qt.Unchecked)

                self.checkBox_drawing.setEnabled(False)
                self.checkBox_drawing.setCheckState(Qt.Unchecked)
                self.checkBox_linked.setEnabled(False)
                self.checkBox_linked.setCheckState(Qt.Checked)
                self.checkBox_vendor.setEnabled(False)
                self.checkBox_vendor.setCheckState(Qt.Unchecked)
# =============================ncr page========================================================
# =============================ncr page========================================================
# =============================ncr page========================================================


# =============================vendors page========================================================
# =============================vendors page========================================================
# =============================vendors page========================================================
    # vendor filter page
    def vendor_filters(self):
            self.vendor_scroll = QScrollArea()
            self.vendor_scroll.setWidgetResizable(True)
            self.vendor_widget = QWidget()

            self.verticalLayout_3 = QVBoxLayout()

            myfont = QFont()
            myfont.setPointSize(12)

            self.label = QLabel("<b>Select VENDOR Filter Options</b>")
            self.label.setFont(myfont)
            self.label.setAlignment(Qt.AlignHCenter)

            self.verticalLayout_3.addWidget(self.label)

            self.groupBox = QGroupBox()

            self.horizontalLayout = QHBoxLayout()
            self.horizontalLayout.setContentsMargins(20, 30, 20, 30)

            self.radioButton_vendor = QRadioButton("Tom H. Filters (default)")
            self.radioButton_vendor.clicked.connect(self.change_vendor_options)
            self.radioButton_vendor.setChecked(True)
            self.radioButton_vendor_custom = QRadioButton("Do Custom Filtering")
            self.radioButton_vendor_custom.clicked.connect(self.change_vendor_options)

            self.horizontalLayout.addStretch()
            self.horizontalLayout.addWidget(self.radioButton_vendor)
            self.horizontalLayout.addWidget(self.radioButton_vendor_custom)
            self.horizontalLayout.addStretch()

            self.groupBox.setLayout(self.horizontalLayout)
            self.verticalLayout_3.addWidget(self.groupBox)
            self.verticalLayout_3.addSpacing(15)

            myfont = QFont()
            myfont.setPointSize(10)

            self.label_2 = QLabel("<b>Vendor Filter Options</b> (for custom filtering)")
            self.label_2.setFont(myfont)
            self.label_2.setAlignment(Qt.AlignHCenter)

            self.verticalLayout_3.addWidget(self.label_2)

            self.line = QFrame()
            self.line.setFrameShape(QFrame.HLine)
            self.line.setFrameShadow(QFrame.Sunken)

            self.verticalLayout_3.addWidget(self.line)
            self.verticalLayout_3.addSpacing(20)

            self.horizontalLayout_2 = QHBoxLayout()

            myfont = QFont()
            myfont.setPointSize(8)

            # blocking this code out for now, may get used for a differnet program
            """
            self.verticalLayout = QVBoxLayout()

            self.label_3 = QLabel("<b>For Singular Vendor Chart</b> (only 1 can be selected at time)")
            self.label_3.setFont(myfont)
            self.label_3.setAlignment(Qt.AlignHCenter)

            self.verticalLayout.addWidget(self.label_3)

            self.listWidget_single_vendor = QListWidget()

            self.verticalLayout.addWidget(self.listWidget_single_vendor)

            self.horizontalLayout_2.addSpacing(50)
            self.horizontalLayout_2.addLayout(self.verticalLayout)
            self.horizontalLayout_2.addSpacing(50)
            """

            self.verticalLayout_2 = QVBoxLayout()

            self.label_4 = QLabel("<b>Vendors To Include/Exclude</b>")
            self.label_4.setFont(myfont)
            self.label_4.setAlignment(Qt.AlignHCenter)

            self.verticalLayout_2.addWidget(self.label_4)

            self.listWidget_vendors = QListWidget()

            self.verticalLayout_2.addWidget(self.listWidget_vendors)

            self.horizontalLayout_2.addStretch()
            self.horizontalLayout_2.addLayout(self.verticalLayout_2)
            self.horizontalLayout_2.addStretch()

            self.verticalLayout_3.addLayout(self.horizontalLayout_2)

            self.vendor_widget.setLayout(self.verticalLayout_3)

            self.vendor_scroll.setWidget(self.vendor_widget)
            self.addWidget(self.vendor_scroll)

            self.set_vendor_options()

    # get vendor options from database
    def set_vendor_options(self):
        # add sqlite3 connection to database for program to get data
        conn = sqlite3.connect("your database")
        cursor = conn.cursor()

        # get all tables with JOB in title and get last one for query
        query = "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE '%QC' ORDER BY name"
        cursor.execute(query)

        # get last ncr table made
        last_ncr_table = "".join(cursor.fetchall()[-1])

        # close sql connection
        conn.close()

        start = time.time()
        # get distinct values from columns
        vendor_query = "SELECT DISTINCT SUPPNAME FROM '" + last_ncr_table + "'" + \
                       "WHERE ITEMCODE GLOB 'AK*' OR ITEMCODE GLOB 'PK*' OR ITEMCODE GLOB 'HK*'"

        # get data frame
        vendor_df = cx.read_sql(connect, vendor_query)
        vendor_df.sort_values(by=["SUPPNAME"], inplace=True)

        # add items to list_vendor
        for i in range(len(vendor_df)):
            item_value = vendor_df.iloc[i].values[0]
            item = QListWidgetItem(item_value)

            # start with items disabled
            item.setFlags(Qt.NoItemFlags)
            item.setCheckState(Qt.Checked)

            self.listWidget_vendors.addItem(item)

     #   selected_items = [self.listWidget_vendors.item(i).text() for i in range(self.listWidget_vendors.count()) if
          #                self.listWidget_vendors.item(i).checkState() == 0]
    #    print(selected_items)

        """
        # blocking this one code for now, may get used some other time for a different program
        # seems limitation with pyqt5 in that i need to create a separate loop for the other qlistwidget
        # add items to list_singular_vendor
        for i in range(len(vendor_df)):
            item_value = vendor_df.iloc[i].values[0]
            item = QListWidgetItem(item_value)
            self.listWidget_single_vendor.addItem(item)

        self.listWidget_single_vendor.setCurrentRow(0)
        """

    # switching vendor options
    def change_vendor_options(self):
        sender = self.sender()

        # update list states based on which radiobutton hit
        if "Custom" in sender.text():
            for index in range(self.listWidget_vendors.count()):
                self.listWidget_vendors.item(index).setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)

        if "Tom" in sender.text():
            for index in range(self.listWidget_vendors.count()):
                self.listWidget_vendors.item(index).setCheckState(Qt.Checked)
                self.listWidget_vendors.item(index).setFlags(Qt.NoItemFlags)
# =============================vendors page========================================================
# =============================vendors page========================================================
# =============================vendors page========================================================
    # about page/description
    def about_page(self):
            self.about_scroll = QScrollArea()
            self.about_scroll.setWidgetResizable(True)
            self.about_widget = QWidget()


            self.about_layout = QVBoxLayout()

            myfont = QFont()
            myfont.setPointSize(12)
            self.about_title = QLabel("<b>About</b>")
            self.about_title.setFont(myfont)
            self.about_title.setAlignment(Qt.AlignHCenter)

            self.about_layout.addWidget(self.about_title)
            self.about_layout.addSpacing(10)

            # Create a QFrame (horizontal line)
            line = QFrame()
            line.setFrameShape(QFrame.HLine)
            line.setFrameShadow(QFrame.Sunken)

            self.about_layout.addWidget(line)
            self.about_layout.addSpacing(15)

            self.about_info = QLabel(
                "<font size='4'>This program is for automatically creating charts based job/ncr/vendor data,<br>"
                "so that it doesn't have to be done manually.  Some things to keep in mind:</font><br><br><br>"
                "     <b>1)</b>  Using custom filters make take a few seconds longer to create the charts<br><br>"
                "     <b>2)</b>  If you add older month data, you will need to restart program to parse that data<br><br>"
                "     <b>3)</b>  The charts will print based on how they look in the GUI, so adjust window size<br><br>"
                "     <b>4)</b>  Don't forget about EpicQuality Export limit of 65,556 lines, you might <br>"
                "                need to export twice and combine into 1 file<br><br>"
                "     <b>5)</b>  When collecting Vendor data from Queries, make sure end date is 1 day after<br><br>"
                "     <b>6)</b>  The Defaults are Tom's settings, with some minor tweaks, such as counting<br>"
                "                 PK99's towards vendor ncr data instead of JOB ncr data.<br><br>"
                "     <b>6)</b>  If you would like any changes or added features, contact Jason Fuller")

            self.about_info.setTextFormat(Qt.RichText)

            self.about_info.setAlignment(Qt.AlignHCenter)
            self.about_layout.addWidget(self.about_info)
            self.about_layout.addStretch()

            self.about_widget.setLayout(self.about_layout)

            self.about_scroll.setWidget(self.about_widget)

            self.addWidget(self.about_scroll)

    # when window is closed, close the sql connection
  #  def close(self):
  #      print("test")
  #      conn.close()

# where data collection happens so UI is still responsive
class DATA(QThread):
    data_changed = pyqtSignal(dict, dict, dict, object, object, object, object, object, object)

    def __init__(self, job_tables, ncr_tables, qc_tables, workcenter_items, process_items, job_items, fa_seq_state,
                 ncr_columns, ncr_drawing, ncr_pks, ncr_novendorpart, vendor_items):
        super().__init__()
        self.job_tables = job_tables
        self.ncr_tables = ncr_tables
        self.qc_tables = qc_tables
        self.workcenter_items = workcenter_items
        self.process_items = process_items
        self.job_items = job_items
        self.fa_seq_state = fa_seq_state
        self.ncr_columns = ncr_columns
        self.ncr_drawing = ncr_drawing
        self.ncr_pks = ncr_pks
        self.ncr_novendorpart = ncr_novendorpart
        self.vendor_items = vendor_items

        # make list of all fa/sa jobs for parsing when collectin ncr data
        self.fa_job_df = []
        self.sa_job_df = []

    def run(self):
        # run functions to collect data
        month_job_data = self.get_job_data()
        ncr_qty_data, fa_ncr_data, sa_ncr_data, pk_ncr_data, ak_ncr_data, total_job_ncr_data, po_ncr_data = self.get_ncr_data()
        vendor_data = self.get_vendor_data()

        # emit data to UI
        self.data_changed.emit(month_job_data, ncr_qty_data, vendor_data, fa_ncr_data, sa_ncr_data,
                               pk_ncr_data, ak_ncr_data, total_job_ncr_data, po_ncr_data)

    # get job data
    def get_job_data(self):

        # base iterative statement i want in the sql select
        statement_jobs = "JOBNO GLOB"

        # complete where statement i want in the sql select
        type_var = " OR ".join(statement_jobs + " '" + i + "*'" for i in self.job_items)

        # job data passed into a dict for each month to create charts off of
        job_dict = {"FA_TOTAL": [], "FA_JOBS": [], "SA_JOBS": [], "AKHK_JOBS": [], "PK_JOBS": [], "TOTAL_JOBS": []}

        start = time.time()
        for i in self.job_tables:
            # table_name = str(''.join(i))
            table = "'" + i + "'"

            query = "SELECT JOBNO, substr(REFID,1, 2) AS REFID, WORKCENTERNAME, PROCESS_ID FROM " + table + " WHERE " + type_var
            df1 = cx.read_sql(connect, query)

            # filter rows out of dataframe
            work_df = df1[~df1['WORKCENTERNAME'].isin(self.workcenter_items)]
            proc_df = work_df[~work_df['PROCESS_ID'].isin(self.process_items)]

            # get fa sequences by var value, which is from the UI
            if self.fa_seq_state == 0:
                # get total FA seq sign offs
                total_fa_seq = (df1["REFID"] == "FA").sum()
            else:
                total_fa_seq = (proc_df["REFID"] == "FA").sum()

            # copy data frame and filter it if necessary
            filtered_df = proc_df.copy()

            # remove any this value
            if self.ncr_pks == 2:
                filtered_df = proc_df[~proc_df["REFID"].str.startswith('PK99')]

            # drop duplicates out of dataframe
            drop_extra = filtered_df.drop_duplicates(subset="JOBNO", keep="first")

            # get all jobs that have FA or SA in them
            get_fa_jobs = drop_extra[drop_extra['REFID'].isin(['FA'])]

            # get fa jobs to list
            append_fa = get_fa_jobs["JOBNO"].to_list()

            # get all jobs that have FA or SA in them
            get_sa_jobs = drop_extra[drop_extra['REFID'].isin(['SA'])]

            # get sa jobs to list
            append_sa = get_sa_jobs["JOBNO"].to_list()

            # append results to main lists
            self.fa_job_df += append_fa
            self.sa_job_df += append_sa

            # get sums for charts and pass to dict
            fa_seq = (drop_extra["REFID"] == "FA").sum()
            sa_seq = (drop_extra["REFID"] == "SA").sum()
            akhk_seq = (drop_extra["REFID"] == "AK").sum() + (drop_extra["REFID"] == "HK").sum()
            pk_seq = (drop_extra["REFID"] == "PK").sum()

            # send to dictionary  to be used for charts
            job_dict["FA_TOTAL"].append(total_fa_seq)
            job_dict["FA_JOBS"].append(fa_seq)
            job_dict["SA_JOBS"].append(sa_seq)
            job_dict["AKHK_JOBS"].append(akhk_seq)
            job_dict["PK_JOBS"].append(pk_seq)
            job_dict["TOTAL_JOBS"].append(pk_seq + fa_seq + sa_seq + akhk_seq)

        end = time.time()
       # print(end - start)

        return job_dict

    # get ncr data
    def get_ncr_data(self):
        # for total ncrs for each category per month for charts
        ncr_dict = {"FA_JOB_NCRS": [], "SA_JOB_NCRS": [], "PK_JOB_NCRS": [], "AK_JOB_NCRS": [], "TOTAL_JOB_NCRS": [],
                    "TOTAL_PO_NCRS": []}

        start = time.time()
        for i in self.ncr_tables:
            table = "'" + i + "'"

            query = "SELECT * FROM " + table + " WHERE Cancelled NOT LIKE 'T%'"

            df1 = cx.read_sql(connect, query)

            # drop any duplicates
            df1 = df1.drop_duplicates(subset="NCR No", keep="first")

            # remove the NaN values
            remove_nas = df1[~df1["Job No/PO"].isna()].astype(str)

            # remove ncr's that aren't written up against drawing
            if "Progress" in remove_nas.head():
                part_number = "Itemcode"
            else:
                part_number = "Part Number"

            # remove rows that don't meet criteria, based on selection in UI
            if self.ncr_drawing == 2:
                # filter to include only these ncr's that have this prefix
                remove_nas = remove_nas[remove_nas[part_number].str.startswith(("AK", "PK", "HK", "RX", "FK"))]

            # separate po/job into different dataframes
            job_ncr = remove_nas[remove_nas["Job No/PO"].str.startswith('JOB')]

            # this is for which column to use if NCR data was extracted out of Enquiries or Epic Quality
            if "Progress" in job_ncr.head():
                if self.ncr_drawing == 0:
                    drawing_column = "Part Number"
                else:
                    drawing_column = "Itemcode"
            else:
                if self.ncr_drawing == 0:
                    drawing_column = "Drawing No"
                else:
                    drawing_column = "Part Number"

            # copy but changed depending on UI selectoin
            filtered_jobs = job_ncr.copy()
            if self.ncr_pks == 2:
                # drop PK99's from JOBS, counted to supplier
                filtered_jobs = job_ncr[~job_ncr[drawing_column].str.startswith('PK99')]

            # get all jobs that have FA or SA in them count number
            get_sas = filtered_jobs[filtered_jobs["Job No/PO"].isin(self.sa_job_df)]
            get_fas = filtered_jobs[filtered_jobs["Job No/PO"].isin(self.fa_job_df)]

            # remove the sa/fa jobs from the jobs dataframes, so that i can get count of remainings
            remove_fa = filtered_jobs[~filtered_jobs["Job No/PO"].isin(self.fa_job_df)]
            remove_sa = remove_fa[~remove_fa["Job No/PO"].isin(self.sa_job_df)]

            # get count of ak/pks
            akhk_ncr = (remove_sa[drawing_column].str.startswith("AK")).sum() + (remove_sa[drawing_column].str.startswith("HK")).sum()
            pk_ncr = (remove_sa[drawing_column].str.startswith("PK")).sum()
            total_jobs = (filtered_jobs["Job No/PO"].str.startswith("JOB")).sum()

            # add to dict
            ncr_dict["SA_JOB_NCRS"].append(len(get_sas))
            ncr_dict["FA_JOB_NCRS"].append(len(get_fas))
            ncr_dict["AK_JOB_NCRS"].append(akhk_ncr)
            ncr_dict["PK_JOB_NCRS"].append(pk_ncr)
            ncr_dict["TOTAL_JOB_NCRS"].append(total_jobs)

            # put po ncrs in a separate column for parsing, include pk99 or not based on var passed through from UI
            if self.ncr_pks == 2:
                po_ncr = remove_nas[remove_nas['Job No/PO'].str.startswith('PO') | remove_nas[drawing_column].str.startswith('PK99')]
            else:
                po_ncr = remove_nas[remove_nas['Job No/PO'].str.startswith('PO')]

            # filter out vendors based on the vendor options from the UI
            if len(self.vendor_items) != 0:
                # remove vendors if removed from list in UI
                po_ncr = po_ncr[~po_ncr["Vendor"].isin(self.vendor_items)]

            # get totals and pass to dict
            total_po = (po_ncr[drawing_column].str.startswith("AK")).sum() + (po_ncr[drawing_column].str.startswith("HK")).sum() + \
                       (po_ncr[drawing_column].str.startswith("PK")).sum()

            # add to ncr dict
            ncr_dict["TOTAL_PO_NCRS"].append(total_po)

            # return all the columns chosen as a dataframe for population into table widget
            if i == self.ncr_tables[-1]:
                # if table has progress in it, change column of data to get
                if "Drawing No" in self.ncr_columns and "Progress" in job_ncr.head():
                    # replace column name based on UI selection
                    if self.ncr_drawing == 2:
                        self.ncr_columns[self.ncr_columns.index("Drawing No")] = "Itemcode"
                    else:
                        self.ncr_columns[self.ncr_columns.index("Drawing No")] = "Part Number"
                else:
                    if self.ncr_drawing == 2:
                        self.ncr_columns[self.ncr_columns.index("Drawing No")] = "Part Number"

                # get new data frame based on columns chosen from UI
                #ncr_data = remove_nas[[x for x in remove_nas.columns if x in self.ncr_columns]]

                # get fa ncr dataframe data, filter columns
                fa_ncr_data = get_fas[[x for x in get_fas.columns if x in self.ncr_columns]]

                # get fa ncr dataframe data, filter columns
                sa_ncr_data = get_sas[[x for x in get_sas.columns if x in self.ncr_columns]]

                # get pk ncr dataframe data & filter columns out
                pk_info = remove_sa[remove_sa[drawing_column].str.startswith("PK")]
                pk_ncr_data = pk_info[[x for x in pk_info.columns if x in self.ncr_columns]]

                # get ak ncr dataframe data & filter columns out
                ak_info = remove_sa[remove_sa[drawing_column].str.startswith("AK") | remove_sa[drawing_column].str.startswith("HK")]
                ak_ncr_data = ak_info[[x for x in ak_info.columns if x in self.ncr_columns]]

                # get total job ncr dataframe data & filter columns out
                total_job_info = remove_sa[remove_sa[drawing_column].str.startswith("AK") | remove_sa[drawing_column].str.startswith("HK") |
                                           remove_sa[drawing_column].str.startswith("PK")]
                total_job_ncr_data = total_job_info[[x for x in total_job_info.columns if x in self.ncr_columns]]

                # get po ncr dataframe data & filter columns out
                po_info = po_ncr[po_ncr[drawing_column].str.startswith("AK") | po_ncr[drawing_column].str.startswith("HK") |
                                 po_ncr[drawing_column].str.startswith("PK")]
                po_ncr_data = po_info[[x for x in po_info.columns if x in self.ncr_columns]]

        end = time.time()
      #  print(end - start)

        return ncr_dict, fa_ncr_data, sa_ncr_data, pk_ncr_data, ak_ncr_data, total_job_ncr_data, po_ncr_data

    # get vendor data
    def get_vendor_data(self):
        # for total ncrs for each category per month for charts
        qc_dict = {"PO_LINE_ITEMS": [], "TOTAL_RECEIVED": []}

        start = time.time()
        for i in self.qc_tables:
            table = "'" + i + "'"

            query = "SELECT ITEMCODE, QTY, SUPPNAME FROM " + table + " WHERE " \
                                                                     "ITEMCODE GLOB 'AK*' OR " \
                                                                     "ITEMCODE GLOB 'PK*' OR " \
                                                                     "ITEMCODE GLOB 'HK*'"

            df1 = cx.read_sql(connect, query)

            if len(self.vendor_items) != 0:
                # remove vendors if removed from list in UI
                df1 = df1[~df1["SUPPNAME"].isin(self.vendor_items)]

            qc_dict["PO_LINE_ITEMS"].append(len(df1))
            qc_dict["TOTAL_RECEIVED"].append(df1["QTY"].sum())

        end = time.time()
       # print(end - start)

        return qc_dict

# QtableWidge with copy/paste functionality for UI
class TableWithCopy(QTableWidget):
    """
    this class extends QTableWidget
    * supports copying multiple cell's text onto the clipboard
    * formatted specifically to work with multiple-cell paste into programs
      like google sheets, excel, or numbers
      and also copying / pasting into cells by more than 1 at a time
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

    def keyPressEvent(self, event):
        super().keyPressEvent(event)
        if event.key() == Qt.Key.Key_C and (event.modifiers() & Qt.KeyboardModifier.ControlModifier):
            copied_cells = sorted(self.selectedIndexes())

            copy_text = ''
            max_column = copied_cells[-1].column()
            for c in copied_cells:
                copy_text += self.item(c.row(), c.column()).text()
                if c.column() == max_column:
                    copy_text += '\n'
                else:
                    copy_text += '\t'
            QApplication.clipboard().setText(copy_text)

        if event.key() == Qt.Key_C and (event.modifiers() & Qt.ControlModifier):
            self.copied_cells = sorted(self.selectedIndexes())
        elif event.key() == Qt.Key_V and (event.modifiers() & Qt.ControlModifier):
            r = self.currentRow() - self.copied_cells[0].row()
            c = self.currentColumn() - self.copied_cells[0].column()
            for cell in self.copied_cells:
                self.setItem(cell.row() + r, cell.column() + c, QTableWidgetItem(cell.data()))

# do sql caching of default queries to get them loaded into cache of database so that when the user
# runs it, it is faster
class sql_cache(QThread):

    def __init__(self, last_12_items, job_items):
        super().__init__()
        self.last_12_items = last_12_items
        self.job_items = job_items

    def run(self):
        # get months into proper format for parsing sql
        self.all_months = self.get_monthsto_parse()

        # run queries
        self.run_sql_queries(self.all_months)

    def run_sql_queries(self, all_months):
        job_suffixes = [" JOB", " PO", " QC"]

        for suffix in job_suffixes:
            start = time.time()
            for month in all_months:
                table_name = month + suffix

                if suffix == " JOB":
                    # base iterative statement i want in the sql select
                    statement_jobs = "JOBNO GLOB"

                    # complete where statement i want in the sql select
                    type_var = " OR ".join(statement_jobs + " '" + i + "*'" for i in self.job_items)

                    query = "SELECT JOBNO, substr(REFID,1, 2) AS REFID, WORKCENTERNAME, PROCESS_ID FROM '" + table_name + "' WHERE " + type_var

                    df1 = cx.read_sql(connect, query)

                if suffix == " PO":
                    query = "SELECT * FROM '" + table_name + "' WHERE Cancelled NOT LIKE 'T%'"

                    df1 = cx.read_sql(connect, query)

                if suffix == " QC":
                    query = "SELECT ITEMCODE, QTY, SUPPNAME FROM '" + table_name + "' WHERE " \
                                                                             "ITEMCODE GLOB 'AK*' OR " \
                                                                             "ITEMCODE GLOB 'PK*' OR " \
                                                                             "ITEMCODE GLOB 'HK*'"

                    df1 = cx.read_sql(connect, query)

            end = time.time()
         #   print(end-start)

    # get months into proper format for parsing sql
    def get_monthsto_parse(self):

        base_months = []

        # transform list of months to match table names for sql querying
        for i in self.last_12_items:
            # get the month as an integer
            index = [x for x, item in enumerate(calendar.month_name) if item == i.split()[1]]
            base_months.append(i.split()[0] + " " + str(index[0]))

        return base_months


if __name__ == "__main__":
    pass